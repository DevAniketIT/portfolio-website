#!/usr/bin/env python3
"""
Excel Automator
Comprehensive Excel automation toolkit for local businesses.
Services priced between ₹2000-5000 per project.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
import numpy as np
from datetime import datetime, timedelta
import os
from typing import List, Dict, Any, Optional
import logging

class ExcelAutomator:
    def __init__(self):
        """Initialize Excel Automator"""
        self.setup_logging()
        
    def setup_logging(self):
        """Setup logging for Excel operations"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
    
    def create_invoice_template(self, company_info: Dict[str, str], output_file: str):
        """
        Create a professional invoice template
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Template"
        
        # Company header
        ws.merge_cells('A1:F1')
        ws['A1'] = company_info.get('name', 'Company Name')
        ws['A1'].font = Font(size=24, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Company details
        row = 2
        for key, value in company_info.items():
            if key != 'name':
                ws[f'A{row}'] = f"{key.title()}: {value}"
                row += 1
        
        # Invoice header
        ws.merge_cells(f'A{row+2}:F{row+2}')
        ws[f'A{row+2}'] = "INVOICE"
        ws[f'A{row+2}'].font = Font(size=18, bold=True)
        ws[f'A{row+2}'].alignment = Alignment(horizontal='center')
        
        # Invoice details section
        start_row = row + 4
        ws[f'A{start_row}'] = "Invoice #:"
        ws[f'A{start_row+1}'] = "Date:"
        ws[f'A{start_row+2}'] = "Due Date:"
        
        ws[f'D{start_row}'] = "Bill To:"
        ws[f'D{start_row+1}'] = "Client Name"
        ws[f'D{start_row+2}'] = "Client Address"
        
        # Items table header
        table_start = start_row + 5
        headers = ['Description', 'Quantity', 'Rate', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Add formulas for amount calculation
        for row_num in range(table_start + 1, table_start + 11):
            ws[f'D{row_num}'] = f'=B{row_num}*C{row_num}'
        
        # Total section
        total_row = table_start + 12
        ws[f'C{total_row}'] = "Subtotal:"
        ws[f'D{total_row}'] = f'=SUM(D{table_start+1}:D{table_start+10})'
        
        ws[f'C{total_row+1}'] = "Tax (%):"
        ws[f'D{total_row+1}'] = f'=D{total_row}*0.18'  # 18% tax
        
        ws[f'C{total_row+2}'] = "TOTAL:"
        ws[f'D{total_row+2}'] = f'=D{total_row}+D{total_row+1}'
        ws[f'C{total_row+2}'].font = Font(bold=True)
        ws[f'D{total_row+2}'].font = Font(bold=True)
        
        # Format currency columns
        for row_num in range(table_start, total_row + 3):
            ws[f'C{row_num}'].number_format = '₹#,##0.00'
            ws[f'D{row_num}'].number_format = '₹#,##0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        self.logger.info(f"Invoice template created: {output_file}")
    
    def create_expense_tracker(self, output_file: str):
        """
        Create an expense tracking spreadsheet with analysis
        """
        wb = openpyxl.Workbook()
        
        # Main expense sheet
        ws_expenses = wb.active
        ws_expenses.title = "Expenses"
        
        # Headers
        headers = ['Date', 'Category', 'Description', 'Amount', 'Payment Method', 'Receipt']
        for col, header in enumerate(headers, 1):
            cell = ws_expenses.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Sample data
        sample_data = [
            [datetime.now().strftime('%Y-%m-%d'), 'Office', 'Stationery', 500, 'Cash', 'Yes'],
            [datetime.now().strftime('%Y-%m-%d'), 'Travel', 'Fuel', 2000, 'Card', 'Yes'],
            [datetime.now().strftime('%Y-%m-%d'), 'Meals', 'Client Lunch', 1500, 'Card', 'No'],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_expenses.cell(row=row_idx, column=col_idx, value=value)
        
        # Format amount column
        for row in range(2, 100):
            ws_expenses[f'D{row}'].number_format = '₹#,##0.00'
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Monthly summary
        ws_summary['A1'] = "Monthly Expense Summary"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = "Category"
        ws_summary['B3'] = "Total Amount"
        
        categories = ['Office', 'Travel', 'Meals', 'Utilities', 'Marketing']
        for idx, category in enumerate(categories, 4):
            ws_summary[f'A{idx}'] = category
            ws_summary[f'B{idx}'] = f'=SUMIF(Expenses.B:B,"{category}",Expenses.D:D)'
            ws_summary[f'B{idx}'].number_format = '₹#,##0.00'
        
        # Create chart
        chart = PieChart()
        chart.title = "Expenses by Category"
        
        data = Reference(ws_summary, min_col=2, min_row=3, max_row=8)
        categories_ref = Reference(ws_summary, min_col=1, min_row=4, max_row=8)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories_ref)
        
        ws_summary.add_chart(chart, "D3")
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        self.logger.info(f"Expense tracker created: {output_file}")
    
    def create_inventory_system(self, output_file: str):
        """
        Create an inventory management system
        """
        wb = openpyxl.Workbook()
        
        # Inventory sheet
        ws_inventory = wb.active
        ws_inventory.title = "Inventory"
        
        headers = ['Item Code', 'Product Name', 'Category', 'Quantity', 'Unit Price', 
                  'Total Value', 'Reorder Level', 'Supplier', 'Last Updated']
        
        for col, header in enumerate(headers, 1):
            cell = ws_inventory.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Add formulas
        for row in range(2, 100):
            ws_inventory[f'F{row}'] = f'=D{row}*E{row}'  # Total Value
            ws_inventory[f'I{row}'] = datetime.now().strftime('%Y-%m-%d')  # Last Updated
        
        # Format currency columns
        for row in range(1, 100):
            ws_inventory[f'E{row}'].number_format = '₹#,##0.00'
            ws_inventory[f'F{row}'].number_format = '₹#,##0.00'
        
        # Low stock alerts
        ws_alerts = wb.create_sheet("Low Stock Alerts")
        ws_alerts['A1'] = "Low Stock Items"
        ws_alerts['A1'].font = Font(size=16, bold=True, color='FF0000')
        
        alert_headers = ['Item Code', 'Product Name', 'Current Qty', 'Reorder Level']
        for col, header in enumerate(alert_headers, 1):
            ws_alerts.cell(row=2, column=col, value=header).font = Font(bold=True)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "Inventory Summary"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = "Total Items:"
        ws_summary['B3'] = "=COUNTA(Inventory.A:A)-1"
        
        ws_summary['A4'] = "Total Inventory Value:"
        ws_summary['B4'] = "=SUM(Inventory.F:F)"
        ws_summary['B4'].number_format = '₹#,##0.00'
        
        ws_summary['A5'] = "Items Below Reorder Level:"
        ws_summary['B5'] = "=SUMPRODUCT((Inventory.D2:D100<Inventory.G2:G100)*(Inventory.D2:D100>0))"
        
        wb.save(output_file)
        self.logger.info(f"Inventory system created: {output_file}")
    
    def create_sales_dashboard(self, data_file: str, output_file: str):
        """
        Create a sales dashboard from existing data
        """
        # Read sales data
        if data_file.endswith('.csv'):
            df = pd.read_csv(data_file)
        else:
            df = pd.read_excel(data_file)
        
        # Create workbook
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
        workbook = writer.book
        
        # Define formats
        title_format = workbook.add_format({'font_size': 16, 'bold': True, 'align': 'center'})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white'})
        currency_format = workbook.add_format({'num_format': '₹#,##0.00'})
        
        # Raw data sheet
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        worksheet_raw = writer.sheets['Raw Data']
        worksheet_raw.set_column('A:Z', 15)
        
        # Dashboard sheet
        worksheet_dash = workbook.add_worksheet('Dashboard')
        
        # Title
        worksheet_dash.merge_range('A1:H1', 'Sales Dashboard', title_format)
        
        # Key metrics
        if 'amount' in df.columns or 'Amount' in df.columns:
            amount_col = 'amount' if 'amount' in df.columns else 'Amount'
            total_sales = df[amount_col].sum()
            avg_sale = df[amount_col].mean()
            
            worksheet_dash.write('A3', 'Total Sales:', header_format)
            worksheet_dash.write('B3', total_sales, currency_format)
            
            worksheet_dash.write('A4', 'Average Sale:', header_format)
            worksheet_dash.write('B4', avg_sale, currency_format)
            
            worksheet_dash.write('A5', 'Number of Sales:', header_format)
            worksheet_dash.write('B5', len(df))
        
        # Monthly sales chart (if date column exists)
        date_columns = [col for col in df.columns if 'date' in col.lower()]
        if date_columns:
            df[date_columns[0]] = pd.to_datetime(df[date_columns[0]])
            monthly_sales = df.groupby(df[date_columns[0]].dt.to_period('M')).sum()
            
            # Write monthly data
            worksheet_dash.write('D3', 'Month', header_format)
            worksheet_dash.write('E3', 'Sales', header_format)
            
            for idx, (month, sales) in enumerate(monthly_sales.iterrows(), 4):
                worksheet_dash.write(f'D{idx}', str(month))
                if amount_col in monthly_sales.columns:
                    worksheet_dash.write(f'E{idx}', sales[amount_col], currency_format)
            
            # Create chart
            chart = workbook.add_chart({'type': 'line'})
            chart.add_series({
                'categories': f'Dashboard!D4:D{idx}',
                'values': f'Dashboard!E4:E{idx}',
                'name': 'Monthly Sales'
            })
            chart.set_title({'name': 'Monthly Sales Trend'})
            chart.set_x_axis({'name': 'Month'})
            chart.set_y_axis({'name': 'Sales Amount'})
            worksheet_dash.insert_chart('G3', chart)
        
        writer.close()
        self.logger.info(f"Sales dashboard created: {output_file}")
    
    def automate_data_entry(self, template_file: str, data_source: str, output_file: str):
        """
        Automate data entry from CSV/Excel into template
        """
        # Load template
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
        
        # Load data
        if data_source.endswith('.csv'):
            df = pd.read_csv(data_source)
        else:
            df = pd.read_excel(data_source)
        
        # Find data start position (first empty row after row 1)
        start_row = 2
        for row in range(2, ws.max_row + 2):
            if all(ws[f'{chr(65+col)}{row}'].value is None for col in range(len(df.columns))):
                start_row = row
                break
        
        # Insert data
        for idx, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + idx, column=col_idx + 1, value=value)
        
        wb.save(output_file)
        self.logger.info(f"Data entry completed: {output_file}")
    
    def create_payroll_system(self, employee_data: List[Dict], output_file: str):
        """
        Create a payroll calculation system
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll"
        
        # Headers
        headers = ['Employee ID', 'Name', 'Basic Salary', 'HRA', 'DA', 'Gross Salary',
                  'PF Deduction', 'Tax Deduction', 'Net Salary']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        
        # Add employee data
        for row_idx, emp in enumerate(employee_data, 2):
            ws[f'A{row_idx}'] = emp.get('id', '')
            ws[f'B{row_idx}'] = emp.get('name', '')
            ws[f'C{row_idx}'] = emp.get('basic_salary', 0)
            
            # Calculate allowances (HRA = 40% of basic, DA = 10% of basic)
            ws[f'D{row_idx}'] = f'=C{row_idx}*0.4'  # HRA
            ws[f'E{row_idx}'] = f'=C{row_idx}*0.1'  # DA
            ws[f'F{row_idx}'] = f'=C{row_idx}+D{row_idx}+E{row_idx}'  # Gross
            
            # Calculate deductions (PF = 12% of basic, Tax based on salary slab)
            ws[f'G{row_idx}'] = f'=C{row_idx}*0.12'  # PF
            ws[f'H{row_idx}'] = f'=IF(F{row_idx}>250000,F{row_idx}*0.05,0)'  # Tax
            ws[f'I{row_idx}'] = f'=F{row_idx}-G{row_idx}-H{row_idx}'  # Net Salary
        
        # Format currency columns
        currency_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
        for col in currency_cols:
            for row in range(2, len(employee_data) + 2):
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
        
        # Summary
        summary_row = len(employee_data) + 4
        ws[f'A{summary_row}'] = "TOTALS:"
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        for col in currency_cols:
            ws[f'{col}{summary_row}'] = f'=SUM({col}2:{col}{len(employee_data)+1})'
            ws[f'{col}{summary_row}'].number_format = '₹#,##0.00'
            ws[f'{col}{summary_row}'].font = Font(bold=True)
        
        wb.save(output_file)
        self.logger.info(f"Payroll system created: {output_file}")

def main():
    automator = ExcelAutomator()
    
    print("Excel Automator Services:")
    print("1. Invoice Template")
    print("2. Expense Tracker")
    print("3. Inventory System")
    print("4. Sales Dashboard")
    print("5. Data Entry Automation")
    print("6. Payroll System")
    
    # Example company info
    company_info = {
        'name': 'Your Company Name',
        'address': '123 Business Street, City',
        'phone': '+91-9876543210',
        'email': 'info@yourcompany.com',
        'gst': 'GST123456789'
    }
    
    # Example usage
    # automator.create_invoice_template(company_info, 'invoice_template.xlsx')
    # automator.create_expense_tracker('expense_tracker.xlsx')
    # automator.create_inventory_system('inventory_system.xlsx')

if __name__ == "__main__":
    main()
